import json
import pandas as pd
from openpyxl import load_workbook

wb = load_workbook(filename="target.xlsx")
ws = wb.active
ws_salsify = wb["inv"]

products = {}

# Using the values_only because you want to return the cells' values
for row in ws.iter_rows(min_row=2,
                        max_row=2000,
                        min_col=1,
                        max_col=12,
                        values_only=True):
    product_id = row[0]
    product = {
        "Title": row[3],
        "Item Type": row[4],
        "Reason": row[11],
    }
    products[product_id] = product

# Using json here to be able to format the output for displaying later
products_json = json.dumps(products, indent=4)
# print(products_json)

## THIS SECTION FILTERS JSON
# Transform json input to python objects
input_dict = json.loads(products_json)

#defines new lists
kill_list = []

# Filter python objects with list comprehensions
iterable_products = list(products.items())
# print(iterable_products)
for id, info in iterable_products:
    #print("\nSKU:", id)
    # for key in info:
    #     print(key + ':', info[key])
    if info["Item Type"] == ("Powered Riding Toys"):
       kill_list.append(id)
    if info["Item Type"] == ("Patio Umbrellas"):
       kill_list.append(id)
    if info["Reason"] == ("Image might be considered to be RACY"):
       kill_list.append(id)
    if info["Reason"] == ("Field may contain suggestive and/or profane language."):
       kill_list.append(id)
    if info["Reason"] == ("Illustrations/Logos are not acceptable images."):
       kill_list.append(id)
    if info["Reason"] == ("Image may be a drawing."):
       kill_list.append(id)
    if info["Reason"] == ("Image might be considered to be ADULT"):
       kill_list.append(id)
    if info["Reason"] == ("Image might be considered to be SPOOF"):
       kill_list.append(id)
    if info["Reason"] == ("Image might be considered to be MEDICAL"):
       kill_list.append(id)
    if info["Reason"] == ("Field may reference a weapon."):
       kill_list.append(id)
    if info["Reason"] == ("Field may reference alcohol."):
       kill_list.append(id)
    if info["Reason"] == ("Field may indicate that the item does not comply with Target's inclusive merchandising policy"):
       kill_list.append(id)

print(kill_list)
print("Length of Kill list:")
print(len(kill_list))

final_list = [x for x in products if x not in kill_list]
print("Length of Product list:")
print(len(products))
print(products)
print(final_list)
print("Length of Final list:")
print(len(final_list))

#Create the spreadsheet!

def add_column(sheet_name, column):
    new_column = ws.max_column + 1

    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=new_column, value=value)

add_column('Sheet3', final_list)
wb.save('output.xlsx')   # use the same name if required